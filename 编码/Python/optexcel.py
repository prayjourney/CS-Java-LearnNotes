# -*- coding:utf-8 -*-
"""
使用openpyxl来操作xlsx格式的excel文档
至于以前的xls格式，暂时不去考虑
openpyxl是python官方推荐的excel处理包
还有xlwings也是一个流行的包，功能较为强大，可以处理VBA
Xlsxwriter样式丰富，但是只能创建，不能修改文档
总体而言，openpyxl是最为综合和全面的一个包
"""

'''
在openpyxl中，主要用到三个概念：Workbooks，Sheets，Cells。
Workbook就是一个excel工作表；Sheet是工作表中的一张表页；Cell就是简单的一个格。
openpyxl就是围绕着这三个概念进行的，
不管读写都是“三板斧”：打开Workbook，定位Sheet，操作Cell。
下面分读和写分别介绍几个常见的方法。
'''

'''
不管是读取还是写入，如果文件是打开的状态，则会报"文件无权限访问"的bug
'''

# 写入
# 关于导入的问题：
# 1.一般而言，一个包之中，类名是不会重复的，所以，我们可以使用["from 包名 import 类名"]的方式来导入
# 2.当一个包里面有子包名和类名重复的时候，我们为了做一个明显的区别，可以使用["from 包名.子包名 import 类名"]的方式来导入
# 更加极端一点，在以上的例子之中我们可以有多个子包名，形成了“包名.子包名.子包名...子包名”的 模式
# 3.还有一种不标准的方式，就是只是并导入子包名，然后使用子包名来引用其中的方法和类，这种方式一般不太推荐
# 但是一般而言，我们都是采用第一种["from 包名 import 类名"]的方式来导入，如下的
# 1.from openpyxl.workbook import Workbook 其实和2.from openpyxl import Workbook等价
# 当然也和3.from openpyxl import workbook等价,但是在下面的使用之中，就要使用wb = workbook.Workbook()的方式来引用
# 而1，2两种方式则只需要使用wb = Workbook()这种模式来引用类之中的方法即可
# 1中的workbook是子包名，Workbook是类名，3之中的workbook也是子包名
# from openpyxl import workbook
# from openpyxl.workbook import Workbook

from openpyxl import Workbook, load_workbook

# 写入
wb = Workbook()  # 创建一个工作簿，一个工作簿创建好之后，默认就有一个sheet
# wb = workbook.Workbook()
sheet = wb.active  # 激活当前的sheet
sheet.title = "默认创建的sheet"
sheet['C3'] = ' hello     world!'
sheet2 = wb.create_sheet("自己创建的sheet2")
sheet2.sheet_properties.tabColor = "1072BA"  # 设置了sheet的属性
sheet2['A3'].value = 'hello'  # 此处设置为A3正确，但是设置为3A错误，说明是先列后行
for x in range(10):
    sheet["A%d" % (x + 1)].value = x + 1

sheet['E1'].value = "=SUM(A:A)"
for st in wb:  # 打印标题
    print(st.title)
wb.save("myfrist.xlsx")  # 保存

# 读取
wr = load_workbook("myfrist.xlsx")
print(wr.sheetnames)
print(sheet['C'])
print(sheet[1][2])  # 从1开始，而非从0开始
print(sheet[1][2].value)  # excel是行优先还是列优先？--->先列后行
print(sheet['C3'].value)  # C3 <-第C3格的值
print(sheet.max_row)  # 10   <-最大行数
print(sheet.max_column)  # 5  <-最大列数
# 打印
for i in sheet['A']:
    print(i.value, end="=")

ss = wr.get_sheet_by_name('自己创建的sheet2')
ss['C4'] = "MLGB"
# 遍历多个单元格
for row in ss.iter_rows('A1:D2'):
    for cell in row:
        print(cell)

wr.save("myfrist.xlsx")
